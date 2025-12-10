import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pymysql
from config import DB_HOST, DB_USER, DB_PASS, DB_NAME
from tkcalendar import DateEntry
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

class RelatorioApp:
    def __init__(self, parent, user):
        self.parent = parent
        self.user = user
        self.frame = ttk.Frame(parent)
        self.frame.pack(fill=tk.BOTH, expand=True)
        
        self.pedidos_data = []
        self.all_pedidos = []
        self.selected_pedido_ids = []

        self.create_widgets()
        self.load_pedidos()

    def create_widgets(self):
        # --- Frame de Filtros ---
        filter_frame = ttk.LabelFrame(self.frame, text="Filtros e Ações")
        filter_frame.pack(padx=10, pady=10, fill=tk.X)

        # Filtro de Pedidos
        ttk.Label(filter_frame, text="Pedidos:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.pedidos_button = ttk.Button(filter_frame, text="Selecionar Pedidos (0)", command=self.open_pedidos_selector)
        self.pedidos_button.grid(row=0, column=1, padx=5, pady=5, sticky='w')

        # Botão de Filtrar
        filter_btn = ttk.Button(filter_frame, text="Filtrar", command=self.filter_data)
        filter_btn.grid(row=0, column=2, padx=5, pady=5)

        # Botões de Exportação
        export_btn = ttk.Button(filter_frame, text="Exportar Resumido", command=lambda: self.export_to_excel(detailed=False))
        export_btn.grid(row=0, column=3, padx=20, pady=5)
        
        export_detailed_btn = ttk.Button(filter_frame, text="Exportar Detalhado", command=lambda: self.export_to_excel(detailed=True))
        export_detailed_btn.grid(row=0, column=4, padx=5, pady=5)

        # --- Frame de Conteúdo (Treeview) ---
        content_frame = ttk.Frame(self.frame)
        content_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        cols = ("Pedido", "Cliente", "Localidade", "Entrega", "Produto/Conjunto", "Lote/OP", "Quantidade", "Status", "Previsão")
        self.tree = ttk.Treeview(content_frame, columns=cols, show="headings")

        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, anchor='w')

        vsb = ttk.Scrollbar(content_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(content_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')
        self.tree.pack(fill=tk.BOTH, expand=True)

    def load_pedidos(self):
        """Carrega a lista de todos os pedidos para o seletor."""
        try:
            connection = self.get_db_connection()
            with connection.cursor() as cursor:
                sql = """
                    SELECT p.idpedido, p.numero_pedido, c.cliente
                    FROM pedido p
                    JOIN add_cliente c ON p.idcliente = c.idcliente
                    ORDER BY p.numero_pedido, c.cliente
                """
                cursor.execute(sql)
                self.all_pedidos = cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar pedidos: {e}")
        finally:
            if connection:
                connection.close()

    def open_pedidos_selector(self):
        """Abre uma janela para selecionar múltiplos pedidos."""
        selector_win = tk.Toplevel(self.frame)
        selector_win.title("Selecionar Pedidos")
        selector_win.geometry("450x500")
        selector_win.transient(self.frame)
        selector_win.grab_set()

        # Frame principal para conter o canvas e o botão
        main_frame = ttk.Frame(selector_win)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame para o canvas e a scrollbar
        canvas_frame = ttk.Frame(main_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        canvas = tk.Canvas(canvas_frame)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Função para habilitar o scroll do mouse
        def _on_mousewheel(event):
            # A velocidade do scroll pode ser ajustada mudando o divisor
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_mousewheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            for child in widget.winfo_children():
                _bind_mousewheel(child)

        self.pedidos_vars = {}
        for pedido in self.all_pedidos:
            var = tk.BooleanVar(value=(pedido['idpedido'] in self.selected_pedido_ids))
            cb = ttk.Checkbutton(scrollable_frame, text=f"{pedido['numero_pedido']} - {pedido['cliente']}", variable=var)
            cb.pack(anchor='w', padx=10, pady=2)
            self.pedidos_vars[pedido['idpedido']] = var
        _bind_mousewheel(scrollable_frame)

        def on_confirm():
            self.selected_pedido_ids = [pid for pid, var in self.pedidos_vars.items() if var.get()]
            self.pedidos_button.config(text=f"Selecionar Pedidos ({len(self.selected_pedido_ids)})")
            selector_win.destroy()

        confirm_btn = ttk.Button(main_frame, text="Confirmar", command=on_confirm)
        confirm_btn.pack(pady=10, padx=10)

    def get_db_connection(self):
        return pymysql.connect(
            host=DB_HOST, user=DB_USER, password=DB_PASS,
            database=DB_NAME, charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )

    def filter_data(self):
        """Busca os dados no banco com base nos filtros e atualiza a Treeview."""
        if not self.selected_pedido_ids:
            messagebox.showwarning("Aviso", "Selecione pelo menos um pedido para filtrar.")
            return

        try:
            connection = self.get_db_connection()
            with connection.cursor() as cursor:
                placeholders = ','.join(['%s'] * len(self.selected_pedido_ids))
                sql = f"""
                    SELECT
                        p.idpedido, p.numero_pedido, c.cliente AS nome_cliente,
                        c.endereco AS endereco_cliente, p.data_entrega,
                        ci.id_item AS id_vinculo, ci.obs_producao, ci.data_prog_fim, ci.quantidade_prod, ci.lote,
                        parent.codigo AS codigo_equipamento,
                        parent.descricao AS nome_equipamento, 
                        child.codigo AS codigo_conjunto, child.descricao AS conjunto,
                        ci.data_engenharia, ci.data_programacao, ci.data_pcp,
                        ci.data_producao, ci.data_qualidade
                    FROM pedido p
                    LEFT JOIN add_cliente c ON p.idcliente = c.idcliente
                    LEFT JOIN cliente_item ci ON p.idpedido = ci.idpedido
                    LEFT JOIN item_composicao ic ON ci.id_composicao = ic.id
                    LEFT JOIN itens parent ON ic.id_item_pai = parent.id
                    LEFT JOIN itens child ON ic.id_item_filho = child.id
                    WHERE p.idpedido IN ({placeholders})
                    ORDER BY p.numero_pedido, c.cliente, parent.descricao, ci.id_item
                """
                cursor.execute(sql, self.selected_pedido_ids)
                results = cursor.fetchall()
                
                # Processar dados e agrupar
                self.pedidos_data = self.process_fetched_data(results)
                self.update_treeview()

        except Exception as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao buscar dados: {e}")
        finally:
            if connection:
                connection.close()

    def process_fetched_data(self, results):
        """Agrupa os resultados da query por pedido."""
        out = {}
        for row in results:
            pedido_id = row['idpedido']
            if pedido_id not in out:
                out[pedido_id] = {
                    'idpedido': pedido_id,
                    'numero_pedido': row['numero_pedido'],
                    'nome_cliente': row['nome_cliente'],
                    'endereco_cliente': row['endereco_cliente'],
                    'data_entrega': row['data_entrega'],
                    'produtos': []
                }
            
            # Derivar status
            status = 'Pendente'
            if row.get('data_qualidade') and row['data_qualidade'].year > 1: status = 'Qualidade'
            elif row.get('data_producao') and row['data_producao'].year > 1: status = 'Produção'
            elif row.get('data_pcp') and row['data_pcp'].year > 1: status = 'PCP'
            elif row.get('data_programacao') and row['data_programacao'].year > 1: status = 'Programação'
            elif row.get('data_engenharia') and row['data_engenharia'].year > 1: status = 'Engenharia'
            
            row['status_producao'] = status
            out[pedido_id]['produtos'].append(row)
            
        return list(out.values())

    def update_treeview(self):
        """Limpa e preenche a Treeview com os dados filtrados."""
        for i in self.tree.get_children():
            self.tree.delete(i)

        for pedido in self.pedidos_data:
            entrega_dt = pedido['data_entrega']
            entrega_str = entrega_dt.strftime('%d/%m/%Y') if entrega_dt else "N/A"
            
            pedido_id = self.tree.insert("", "end", values=(
                pedido['numero_pedido'],
                pedido['nome_cliente'],
                pedido['endereco_cliente'], 
                entrega_str,
                "", "", ""
            ))
            self.tree.item(pedido_id, tags=('pedido_header',))

            equip_anterior = None
            for produto in pedido['produtos']:
                if produto['nome_equipamento'] != equip_anterior:
                    equip_anterior = produto['nome_equipamento']
                    codigo_equip = produto.get('codigo_equipamento', '')
                    display_equip = f"{codigo_equip} - {equip_anterior}" if codigo_equip else equip_anterior
                    equip_id = self.tree.insert(pedido_id, "end", values=(
                        "", "", "", "", f"Equip: {display_equip}", "", "", "", ""
                    ))
                    self.tree.item(equip_id, tags=('equip_header',))

                previsao_dt = produto['data_prog_fim']
                previsao_str = previsao_dt.strftime('%d/%m/%Y') if previsao_dt else "N/A"
                codigo_conjunto = produto.get('codigo_conjunto', '')
                display_conjunto = f"{codigo_conjunto} - {produto['conjunto']}" if codigo_conjunto and produto['conjunto'] else produto['conjunto'] or "N/A"
                
                self.tree.insert(equip_id, "end", values=(
                    "", "", "", "",
                    display_conjunto,
                    produto.get('lote', ''),
                    produto.get('quantidade_prod', ''),
                    produto['status_producao'],
                    previsao_str
                ))

        self.tree.tag_configure('pedido_header', background='#E0E0E0', font=('Arial', 10, 'bold'))
        self.tree.tag_configure('equip_header', background='#F0F0F0', font=('Arial', 9, 'italic'))

    def export_to_excel(self, detailed=False):
        if not self.pedidos_data:
            messagebox.showwarning("Aviso", "Não há dados para exportar. Por favor, filtre os pedidos primeiro.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Salvar Relatório Como"
        )
        if not filepath:
            return

        wb = Workbook()
        ws = wb.active

        if detailed:
            self.create_detailed_report(ws)
        else:
            self.create_summary_report(ws)

        try:
            wb.save(filepath)
            messagebox.showinfo("Sucesso", f"Relatório salvo com sucesso em:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar o arquivo: {e}")

    def create_summary_report(self, ws):
        ws.title = "Relatorio Resumido"
        headers = ['Pedido', 'Cliente', 'Previsão de Término de Produção', 'Produto', 'Lote/OP', 'Quantidade', 'Status', 'Observação', 'Previsão de Entrega']
        ws.append(headers)
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        row_idx = 2
        for pedido in self.pedidos_data:
            # Lógica de agrupamento de produtos similar ao PHP
            # (Simplificado para o exemplo, pode ser expandido)
            start_row = row_idx
            for produto in pedido['produtos']:
                entrega_dt = pedido['data_entrega']
                entrega_str = entrega_dt.strftime('%d/%m/%Y') if entrega_dt else "N/A"
                previsao_dt = produto['data_prog_fim']
                previsao_str = previsao_dt.strftime('%d/%m/%Y') if previsao_dt else "N/A"

                codigo_equip = produto.get('codigo_equipamento', '')
                display_equip = f"{codigo_equip} - {produto['nome_equipamento']}" if codigo_equip else produto['nome_equipamento']

                ws.cell(row=row_idx, column=1, value=pedido['numero_pedido'])
                ws.cell(row=row_idx, column=2, value=f"{pedido['nome_cliente']} - {pedido['endereco_cliente']}")
                ws.cell(row=row_idx, column=3, value=entrega_str)
                ws.cell(row=row_idx, column=4, value=display_equip)
                ws.cell(row=row_idx, column=5, value=produto.get('lote', ''))
                ws.cell(row=row_idx, column=6, value=produto.get('quantidade_prod', ''))
                ws.cell(row=row_idx, column=7, value=produto['status_producao'])
                ws.cell(row=row_idx, column=8, value=produto['obs_producao'])
                ws.cell(row=row_idx, column=9, value=previsao_str)
                
                for col in range(1, 10):
                    ws.cell(row=row_idx, column=col).alignment = center_align
                    ws.cell(row=row_idx, column=col).border = border
                row_idx += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def create_detailed_report(self, ws):
        ws.title = "Pedidos em Carteira"
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Título principal
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'PEDIDOS EM CARTEIRA'
        title_cell.font = title_font
        title_cell.fill = header_fill
        title_cell.alignment = center_align

        row_idx = 2
        for pedido in self.pedidos_data:
            # Cabeçalho do Pedido
            ws.cell(row=row_idx, column=1, value='Pedido').font = header_font; ws.cell(row=row_idx, column=1).fill = header_fill
            ws.cell(row=row_idx, column=2, value=pedido['numero_pedido'])
            ws.cell(row=row_idx, column=3, value='Cliente').font = header_font; ws.cell(row=row_idx, column=3).fill = header_fill
            ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=8)
            ws.cell(row=row_idx, column=4, value=pedido['nome_cliente'])
            row_idx += 1
            
            ws.cell(row=row_idx, column=1, value='Localidade').font = header_font; ws.cell(row=row_idx, column=1).fill = header_fill
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
            ws.cell(row=row_idx, column=2, value=pedido['endereco_cliente'])
            row_idx += 1

            entrega_dt = pedido['data_entrega']
            entrega_str = entrega_dt.strftime('%d/%m/%Y') if entrega_dt else "N/A"
            ws.cell(row=row_idx, column=1, value='Previsão de Término de Produção').font = header_font; ws.cell(row=row_idx, column=1).fill = header_fill
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
            ws.cell(row=row_idx, column=2, value=entrega_str)
            row_idx += 1

            # Cabeçalho dos Produtos
            prod_headers = ['Produto', 'Conjunto', 'Lote/OP', 'Quantidade', 'Data Engenharia', 'Status Produção', 'Observação Produção', 'Previsão Entrega']
            ws.append([]) # Linha em branco
            header_row = ws.row_dimensions[row_idx]
            header_row.height = 20
            for i, h in enumerate(prod_headers, 1):
                cell = ws.cell(row=row_idx, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            row_idx += 1

            # Identificar todos os itens que são "filhos" (conjuntos) para não listá-los como "pais"
            itens_filhos = {p['conjunto'] for p in pedido['produtos'] if p['conjunto']}

            # Adicionar apenas os produtos que não são filhos de outros itens
            for produto in pedido['produtos']:
                # A condição principal: só adiciona a linha se o 'nome_equipamento' não for um 'conjunto'
                if produto['nome_equipamento'] not in itens_filhos:
                    eng_dt = produto['data_engenharia']
                    eng_str = eng_dt.strftime('%d/%m/%Y') if eng_dt else ""
                    previsao_dt = produto['data_prog_fim']
                    previsao_str = previsao_dt.strftime('%d/%m/%Y') if previsao_dt else 'N/A'

                    codigo_equip = produto.get('codigo_equipamento', '')
                    display_equip = f"{codigo_equip} - {produto['nome_equipamento']}" if codigo_equip else produto['nome_equipamento']
                    
                    codigo_conjunto = produto.get('codigo_conjunto', '')
                    display_conjunto = f"{codigo_conjunto} - {produto['conjunto']}" if codigo_conjunto and produto['conjunto'] else produto['conjunto']

                    data_row = [
                        display_equip,
                        display_conjunto,
                        produto.get('lote', ''),
                        produto.get('quantidade_prod', ''),
                        eng_str,
                        produto['status_producao'],
                        produto['obs_producao'],
                        previsao_str
                    ]
                    ws.append(data_row)
                    for col in range(1, 9):
                        ws.cell(row=row_idx, column=col).alignment = center_align
                        ws.cell(row=row_idx, column=col).border = border
                    row_idx += 1
            
            row_idx += 1 # Espaço entre pedidos
        
        # Auto-ajuste das colunas
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
